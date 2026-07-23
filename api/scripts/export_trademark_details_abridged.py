"""Sanity-check trademark unique URLs and export abridged detail workbook."""
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from sqlalchemy import create_engine, text

API_ROOT = Path(__file__).resolve().parents[1]
ROOT = API_ROOT.parent
sys.path.insert(0, str(API_ROOT))

from app.config import get_settings  # noqa: E402

CATALOG_DIR = ROOT / "docs" / "Domains" / "Trademarks"
OUT_PATH = CATALOG_DIR / "Trademark_Source_Details_Abridged_Export.xlsx"

ABRIDGED_HEADERS = [
    "catalog_id",
    "country",
    "country_code",
    "jurisdiction",
    "office",
    "search_url",
    "status_lookup_url",
    "filing_url",
    "gazette_url",
    "api_url",
    "api_docs_url",
    "bulk_download_url",
    "access_type",
    "authentication",
    "rate_limit",
    "supports_nice_classes",
    "supports_image_search",
    "update_frequency",
    "status",
    "last_verified",
    "notes",
]

# Thin catalog for category / description / source_type hints
HEADER_ALIASES = {
    "catalog id": "catalog_id",
    "catalog_id": "catalog_id",
    "name": "name",
    "source url": "source_url",
    "source_url": "source_url",
    "category": "category",
    "description": "description",
    "source type": "source_type",
    "source_type": "source_type",
}


def _norm_header(v: object) -> str | None:
    if v is None:
        return None
    return HEADER_ALIASES.get(str(v).strip().lower())


def _cell(v: object) -> str:
    return "" if v is None else str(v).strip()


def load_catalog() -> dict[str, dict[str, str]]:
    rows: dict[str, dict[str, str]] = {}
    files = sorted(
        [
            *CATALOG_DIR.glob("Trademark_Intelligence_Source_Catalog_Batch*.csv"),
            *CATALOG_DIR.glob("Trademark_Intelligence_Source_Catalog_Batch*.xlsx"),
        ],
        key=lambda p: p.name.lower(),
    )
    for path in files:
        if path.name.startswith(".~"):
            continue
        batch: list[dict[str, str]] = []
        if path.suffix.lower() == ".csv":
            with path.open(encoding="utf-8-sig", newline="") as f:
                for raw in csv.DictReader(f):
                    row: dict[str, str] = {}
                    for k, v in raw.items():
                        nk = _norm_header(k)
                        if nk:
                            row[nk] = _cell(v)
                    batch.append(row)
        else:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            header = [_norm_header(h) for h in next(it)]
            for values in it:
                if not values or all(v is None or str(v).strip() == "" for v in values):
                    continue
                row = {}
                for key, value in zip(header, values):
                    if key:
                        row[key] = _cell(value)
                batch.append(row)
            wb.close()
        for row in batch:
            cid = (row.get("catalog_id") or "").strip().upper()
            if cid:
                rows[cid] = row
    return rows


def infer_country_code(name: str, category: str, url: str) -> str:
    text = f"{name} {category} {url}".lower()
    mapping = [
        ("uspto", "US"),
        ("united states", "US"),
        ("wipo", "WIPO"),
        ("euipo", "EU"),
        ("tmdn", "EU"),
        ("tmview", "EU"),
        ("boip", "BX"),
        ("benelux", "BX"),
        ("aripo", "ARIPO"),
        ("oapi", "OAPI"),
        ("canada", "CA"),
        ("cipo", "CA"),
        ("uk ", "GB"),
        ("ipo.gov.uk", "GB"),
        ("australia", "AU"),
        ("ipaustralia", "AU"),
        ("new zealand", "NZ"),
        ("japan", "JP"),
        ("j-platpat", "JP"),
        ("china", "CN"),
        ("cnipa", "CN"),
        ("korea", "KR"),
        ("kipris", "KR"),
        ("india", "IN"),
        ("ipindia", "IN"),
        ("brazil", "BR"),
        ("inpi.gov.br", "BR"),
        ("mexico", "MX"),
        ("germany", "DE"),
        ("dpma", "DE"),
        ("france", "FR"),
        ("switzerland", "CH"),
        ("swissreg", "CH"),
        ("spain", "ES"),
        ("oepm", "ES"),
        ("italy", "IT"),
        ("portugal", "PT"),
        ("ireland", "IE"),
        ("sweden", "SE"),
        ("norway", "NO"),
        ("denmark", "DK"),
        ("finland", "FI"),
        ("poland", "PL"),
        ("czech", "CZ"),
        ("slovakia", "SK"),
        ("hungary", "HU"),
        ("romania", "RO"),
        ("croatia", "HR"),
        ("slovenia", "SI"),
        ("türkiye", "TR"),
        ("turkey", "TR"),
        ("turkpatent", "TR"),
        ("greece", "GR"),
        ("saudi", "SA"),
        ("saip", "SA"),
        ("uae", "AE"),
        ("israel", "IL"),
        ("qatar", "QA"),
        ("bahrain", "BH"),
        ("thailand", "TH"),
        ("philippines", "PH"),
        ("ipophil", "PH"),
        ("singapore", "SG"),
        ("ipos", "SG"),
        ("malaysia", "MY"),
        ("myipo", "MY"),
        ("indonesia", "ID"),
        ("vietnam", "VN"),
        ("hong kong", "HK"),
        ("taiwan", "TW"),
        ("south africa", "ZA"),
        ("cipc", "ZA"),
        ("kenya", "KE"),
        ("nigeria", "NG"),
        ("ghana", "GH"),
        ("egypt", "EG"),
        ("morocco", "MA"),
        ("tunisia", "TN"),
        ("georgia", "GE"),
        ("sakpatenti", "GE"),
        ("ukraine", "UA"),
        ("russia", "RU"),
        ("fips", "RU"),
        ("asean", "ASEAN"),
    ]
    for needle, code in mapping:
        if needle in text:
            return code
    # US state registries
    if "state" in category.lower() or "secretary of state" in text or "sos." in text:
        return "US-ST"
    return ""


def infer_jurisdiction(category: str, country_code: str) -> str:
    cat = (category or "").lower()
    if "state" in cat:
        return "state"
    if country_code in {"WIPO", "ARIPO", "OAPI", "ASEAN", "EU", "BX"}:
        return "regional" if country_code != "WIPO" else "international"
    if "court" in cat or "ttab" in cat:
        return "court"
    if "global" in cat or "international" in cat:
        return "international"
    if "regional" in cat:
        return "regional"
    return "national"


def infer_access_type(source_type: str, category: str, name: str) -> str:
    st = (source_type or "").lower()
    blob = f"{category} {name}".lower()
    if "rest_api" in st or "api" in blob and "search" not in blob:
        return "rest_api"
    if "bulk" in st or "bulk" in blob:
        return "bulk_download"
    if "gazette" in blob:
        return "gazette"
    if "search" in blob or "registry" in blob:
        return "search"
    return "portal"


def infer_office(name: str) -> str:
    # Use name as office label for now (catalog names are office-ish).
    return (name or "").strip()[:256]


def main() -> None:
    engine = create_engine(get_settings().database_url_sync)
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT catalog_id, name, source_url, category, description, status::text
                FROM sources
                WHERE domain = 'trademarks'
                  AND catalog_id IS NOT NULL
                ORDER BY catalog_id
                """
            )
        ).fetchall()
        distinct = conn.execute(
            text(
                """
                SELECT COUNT(DISTINCT source_url)
                FROM sources
                WHERE domain = 'trademarks'
                  AND source_url IS NOT NULL
                  AND source_url <> ''
                """
            )
        ).scalar()
        dup_urls = conn.execute(
            text(
                """
                SELECT source_url, COUNT(*) AS n
                FROM sources
                WHERE domain = 'trademarks'
                  AND source_url IS NOT NULL
                  AND source_url <> ''
                GROUP BY source_url
                HAVING COUNT(*) > 1
                ORDER BY n DESC, source_url
                """
            )
        ).fetchall()

    print(f"DB trademark rows: {len(rows)}")
    print(f"Distinct source_url: {distinct}")
    print(f"URLs shared by >1 row: {len(dup_urls)}")
    if len(rows) != distinct:
        print("NOTE: row count != distinct URL count")
        for url, n in dup_urls[:10]:
            print(f"  x{n} {url}")

    catalog = load_catalog()
    wb = Workbook()
    ws = wb.active
    ws.title = "details_abridged"
    ws.append(ABRIDGED_HEADERS)

    for catalog_id, name, source_url, category, description, status in rows:
        cat = catalog.get(catalog_id or "", {})
        source_type = cat.get("source_type") or ""
        category = category or cat.get("category") or ""
        url = (source_url or "").rstrip("/")
        country_code = infer_country_code(name or "", category, url)
        jurisdiction = infer_jurisdiction(category, country_code)
        access_type = infer_access_type(source_type, category, name or "")
        # Leave specialist URL columns blank for enrichment; seed search_url from source_url
        # when category looks like search/registry, else leave blank and put URL in notes? 
        # User wants export of the 266 — put thin source_url into search_url as starting point
        # for portals that are search UIs; still copy to notes pointer.
        row = {
            "catalog_id": catalog_id,
            "country": "",
            "country_code": country_code,
            "jurisdiction": jurisdiction,
            "office": infer_office(name or ""),
            "search_url": url,
            "status_lookup_url": "",
            "filing_url": "",
            "gazette_url": "",
            "api_url": url if access_type == "rest_api" else "",
            "api_docs_url": "",
            "bulk_download_url": url if access_type == "bulk_download" else "",
            "access_type": access_type,
            "authentication": "",
            "rate_limit": "",
            "supports_nice_classes": "",
            "supports_image_search": "",
            "update_frequency": "",
            "status": (status or "active").lower(),
            "last_verified": "",
            "notes": (description or "")[:500],
        }
        # If api/bulk took the URL, still keep search_url for convenience unless pure API
        if access_type == "rest_api":
            row["search_url"] = ""
        if access_type == "bulk_download":
            row["search_url"] = ""
        if access_type == "gazette":
            row["gazette_url"] = url
            row["search_url"] = ""
        ws.append([row[h] for h in ABRIDGED_HEADERS])

    # Summary sheet
    ws2 = wb.create_sheet("sanity")
    ws2.append(["metric", "value"])
    ws2.append(["trademark_rows", len(rows)])
    ws2.append(["distinct_source_url", distinct])
    ws2.append(["duplicate_url_groups", len(dup_urls)])
    ws2.append(["catalog_ids_expected", 320])
    ws2.append(["gap_vs_320", 320 - len(rows)])
    ws2.append(["export_rows", len(rows)])
    ws2.append([])
    ws2.append(["note", "search_url prefilled from sources.source_url (cleared when access_type is api/bulk/gazette)"])

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"Export rows: {len(rows)}")


if __name__ == "__main__":
    main()
